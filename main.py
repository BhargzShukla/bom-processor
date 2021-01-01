"""
Go line by line
Create dataclass
Store previous level and previous row
If level == 1, copy row as is
If level != 1, create new item as level 1
Parse this list into individual sheets for result

Conditions for previous level:
if curr_item.level == 1: no change, copy row as is
if curr_item.level > prev_item.level: no change
if curr_item.level < prev_item.level:
"""

import pandas as pd
import openpyxl
from dataclasses import dataclass
import re


@dataclass
class Item:
    name: str
    level: str
    raw_material: str
    quantity: float
    unit: str

    def __post_init__(self):
        self.level = re.sub(r'[.]', r'', str(self.level))


COL_ITEM_NAME = 0
COL_LEVEL = 1
COL_RAW_MATERIAL = 2
COL_QUANTITY = 3
COL_UNIT = 4

wb = openpyxl.load_workbook('data_test.xlsx')
sheet = wb.active

raw_items = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        raw_items.append(Item(name=row[COL_ITEM_NAME], level=row[COL_LEVEL], raw_material=row[COL_RAW_MATERIAL],
                              quantity=row[COL_QUANTITY], unit=row[COL_UNIT]))

processed_items = []
last_processed = None
for idx, item in enumerate(raw_items):
    prev_item = raw_items[idx - 1] if idx > 0 else item
    prev_level = prev_item.level
    if item.level == '1':
        processed_items.append(item)
        prev_level = item.level
        last_processed = item
    elif item.level > prev_item.level:
        processed_items.append(
            Item(name=prev_item.raw_material, level=last_processed.level, raw_material=item.raw_material,
                 quantity=item.quantity, unit=item.unit))
    elif item.level == prev_item.level:
        processed_items.append(
            Item(name=processed_items[-1].name, level=processed_items[-1].level, raw_material=item.raw_material,
                 quantity=item.quantity, unit=item.unit))


# Some global settings
# Use 3 decimal places in output display
pd.set_option("display.precision", 3)

# Don't wrap repr(DataFrame) across additional lines
pd.set_option("display.expand_frame_repr", False)

# Set max rows displayed in output to 25
pd.set_option("display.max_rows", 25)


def clean_dataset_from_file(data_file: str) -> pd.DataFrame:
    df = pd.read_excel(data_file, engine='openpyxl', keep_default_na=False)
    df = df[df['Item Name'].astype(bool)]
    df.replace(regex=True, to_replace=r'[.]', value='', inplace=True)
    df['Level'] = pd.to_numeric(df['Level'])
    return df


def get_item_groups(dataset: pd.DataFrame):
    return dataset.groupby('name', as_index=False, sort=False).groups


data = pd.DataFrame(processed_items)
item_count = get_item_groups(data)

test_wb = openpyxl.Workbook()

for idx, (key, val) in enumerate(item_count.items(), start=1):
    sheet = wb.create_sheet(title=key)
    sheet['a1'] = 'Finished Good List'
    sheet['a2'] = '#'
    sheet['b2'] = 'Item Description'
    sheet['c2'] = 'Quantity'
    sheet['d2'] = 'Unit'
    sheet['a3'] = '1'
    sheet['b3'] = key
    sheet['c3'] = '1'
    sheet['d3'] = 'Pc'
    sheet['a4'] = 'End of FG'
    sheet['a5'] = 'Raw Material List'
    sheet['a6'] = '#'
    sheet['b6'] = 'Item Description'
    sheet['c6'] = 'Quantity'
    sheet['d6'] = 'Unit'
    for item_idx, item_val in enumerate(list(val), start=1):
        item_name = data.iloc[item_val]['raw_material']
        item_quantity = data.iloc[item_val]['quantity']
        item_unit = data.iloc[item_val]['unit']
        sheet.append([item_idx, item_name, item_quantity, item_unit])
    sheet.append(['End of RM'])

wb.save('data_test.xlsx')
