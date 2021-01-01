from openpyxl import load_workbook

from src.classes import Item
from src.constants import (
    COL_ITEM_NAME,
    COL_LEVEL,
    COL_RAW_MATERIAL,
    COL_QUANTITY,
    COL_UNIT,
)


def load_data_from_file(data_file):
    source_workbook = load_workbook(data_file)
    source_sheet = source_workbook.get_sheet_by_name("Source")
    return source_sheet, source_workbook


def get_data_from_source_sheet(sheet):
    """
    Convert the source spreadsheet data into a list of `Item`s.
    :param sheet:
    :return:
    """
    source_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            item = Item(
                name=row[COL_ITEM_NAME],
                level=row[COL_LEVEL],
                raw_material=row[COL_RAW_MATERIAL],
                quantity=row[COL_QUANTITY],
                unit=row[COL_UNIT],
            )
            source_data.append(item)
    return source_data


def flatten(source):
    """
    Flatten the source dataset. Converts each item into a top level item with the appropriate raw material.
    :param source: The raw item list
    :return: list
    """
    processed_items = []
    last_top_processed = None
    for idx, item in enumerate(source):
        prev_item = source[idx - 1] if idx > 0 else item
        if item.level == "1":
            processed_items.append(item)
            last_top_processed = item
        elif item.level > prev_item.level:
            new_item = Item(
                name=prev_item.raw_material,
                level=last_top_processed.level,
                raw_material=item.raw_material,
                quantity=item.quantity,
                unit=item.unit,
            )
            processed_items.append(new_item)
        elif item.level == prev_item.level:
            new_item = Item(
                name=processed_items[-1].name,
                level=processed_items[-1].level,
                raw_material=item.raw_material,
                quantity=item.quantity,
                unit=item.unit,
            )
            processed_items.append(new_item)
    return processed_items


def get_item_groups(dataset):
    """
    Group items from the dataset according to their `Item Name` value.
    :param dataset: Flattened source dataset.
    :return: dict
    """
    return dataset.groupby("name", as_index=False, sort=False).groups


def custom_sheet(workbook, title, key):
    sheet = workbook.create_sheet(title=title)
    sheet["a1"] = "Finished Good List"
    sheet["a2"] = "#"
    sheet["b2"] = "Item Description"
    sheet["c2"] = "Quantity"
    sheet["d2"] = "Unit"
    sheet["a3"] = 1
    sheet["b3"] = key
    sheet["c3"] = 1
    sheet["d3"] = "Pc"
    sheet["a4"] = "End of FG"
    sheet["a5"] = "Raw Material List"
    sheet["a6"] = "#"
    sheet["b6"] = "Item Description"
    sheet["c6"] = "Quantity"
    sheet["d6"] = "Unit"
    return sheet
